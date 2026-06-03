from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


def parse_xlsx(source: bytes | str | Path) -> list[dict]:
    """Returns list of {page_num=<sheet_idx>, text, kind='sheet'}.

    每个 sheet 一个 chunk；rows 用 ' | ' 分隔，列用 tab 分隔。
    """
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(filename=BytesIO(source), data_only=True, read_only=True)
    else:
        wb = load_workbook(filename=str(source), data_only=True, read_only=True)

    chunks: list[dict] = []
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows_text.append("\t".join(cells))
        body = " | ".join(rows_text)
        if body.strip():
            chunks.append({
                "page_num": i,
                "text": f"[Sheet: {sheet_name}]\n{body}",
                "kind": "sheet",
            })
    wb.close()
    return chunks
